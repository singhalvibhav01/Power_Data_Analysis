# -*- coding: utf-8 -*-
"""
EPA Complete Generator (updated Metrics + existing Wholesale)
============================================================
Metrics workflow (updated):
---------------------------
- Use four Snowflake SQLs:
    * epa_metrics_price
    * epa_metrics_thermal
    * epa_metrics_wind_solar
    * epa_nuclear_flex
- Populate Deliverables/EPA/EPA Metrics.xlsx:
    * Sheet 'Prices'     : columns A-F from epa_metrics_price
    * Sheet 'Thermal_DB' : columns A-F from epa_metrics_thermal
    * Sheet 'W&S_DB'     : columns A-F from epa_metrics_wind_solar
    * Sheet 'Nuclear_DB' : columns A-F from epa_nuclear_flex
- Step 2: Refresh + calculate EPA Metric (or 202506).xlsx.
- Step 3: Refresh + calculate and refresh all workbook links in:
    Deliverables/EPA/EPA - Capacity prices for EPA.xlsx
- Template:
    Deliverables/EPA/S&P Global—EPA April 2025—Planning case—
      Metrics up to 2050 (LINKS).xlsx
    * Copy to timestamped "with links" file
    * Update all workbook links, including gas & EUA vintages (if provided)
    * Refresh + calculate the with-links file
    * Copy and break all links to create a "without links" file

Wholesale workflow:
-------------------
- Unchanged from previous code:
    * Read hourly prices from Snowflake (epa_hourly_prices)
    * Write into interval workbook (raw data!A4)
    * Refresh interval and wholesale xlsb template
    * Update links with gas/EUA/coal vintages
    * Create with-links and without-links xlsb outputs
"""

import shutil
import logging
import re
import os
import stat
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Iterable, Tuple
import pandas as pd
import numpy as np
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.excel_template_manager import ExcelTemplateManager
from extractors.snowflake_connector import SnowflakeConnector, QueryManager

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


class EPAGenerator:
    """
    Unified generator for EPA deliverables (Metrics + Wholesale Prices).
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        template_dir: Optional[str] = None,
        run_id: Optional[str] = None,
        vintage_gas_path: Optional[str] = None,
        vintage_eua_path: Optional[str] = None,
        vintage_coal_path: Optional[str] = None,
        connector: Optional[SnowflakeConnector] = None,
    ):
        # Shared Snowflake connector
        self.connector = connector if connector else SnowflakeConnector()
        self.query_manager = QueryManager()
        self.run_id = run_id

        script_dir = Path(__file__).resolve().parent
        project_root = self._find_project_root(script_dir)
        default_epa_dir = project_root / "Deliverables" / "EPA"

        # Directories
        self.template_dir = Path(template_dir) if template_dir else default_epa_dir
        self.output_dir = Path(output_dir) if output_dir else default_epa_dir
        self.output_dir_generated = self.output_dir / "Generated"
        self.output_dir_generated.mkdir(parents=True, exist_ok=True)

        # Metrics-specific files
        self.metrics_static_file = self.output_dir / "EPA Metrics.xlsx"
        self.capacity_prices_file = self.output_dir / "EPA - Capacity prices for EPA.xlsx"
        self.metrics_template_path = (
            self.template_dir
            / "EPA Metrics - Planning case.xlsx"
        )

        # Wholesale templates
        self.wholesale_template_path = (
            self.template_dir
            / "Wholesale prices - Planning Case.xlsx"
        )
        self.interval_file_path = self.template_dir / "EPA interval main file.xlsx"

        # Vintage paths
        self.vintage_gas_path = Path(vintage_gas_path) if vintage_gas_path else None
        self.vintage_eua_path = Path(vintage_eua_path) if vintage_eua_path else None
        self.vintage_coal_path = Path(vintage_coal_path) if vintage_coal_path else None

        # Cache
        self._data_cache: Dict[str, pd.DataFrame] = {}
        self._currency_pattern = re.compile(r"[\$\€\£\¥₹]")

    # -------------------------------------------------------------------------
    # Common helpers
    # -------------------------------------------------------------------------
    @staticmethod
    def _find_project_root(start: Path) -> Path:
        for p in [start] + list(start.parents):
            if (p / "Deliverables" / "EPA").exists():
                return p
            if (p / "Automated Process" / "Deliverables" / "EPA").exists():
                return p / "Automated Process"
        return start.parents[1] if len(start.parents) > 1 else start.parent

    def _get_run_id(self) -> str:
        if self.run_id is None:
            self.run_id = self.connector.get_current_run_id()
        return self.run_id

    @staticmethod
    def _make_writable(file_path: Path) -> None:
        """Ensure file is writable by stripping any Read-Only attribute."""
        if file_path.exists():
            os.chmod(file_path, stat.S_IWRITE)

    # -------------------------------------------------------------------------
    # METRICS: data loading
    # -------------------------------------------------------------------------
    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        column_mapping = {
            'Power Plant': 'Plant',
            'power plant': 'Plant',
            'POWER PLANT': 'Plant',
            'Market': 'Plant',
            'market': 'Plant',
            'MARKET': 'Plant',
            'YEAR': 'Year',
            'year': 'Year',
            'PROPERTY': 'Property',
            'property': 'Property',
            'VALUE': 'Value',
            'value': 'Value',
            'UNIT': 'Unit',
            'unit': 'Unit',
        }
        return df.rename(columns=column_mapping)

    def _load_query_data(self, query_name: str) -> pd.DataFrame:
        if query_name in self._data_cache:
            return self._data_cache[query_name]
        print(f"  Loading {query_name} from Snowflake...")
        try:
            query_path = self.query_manager.get_query_path(query_name)
            if not query_path.exists():
                logger.warning(f"Query not found: {query_name}")
                return pd.DataFrame()
            df = self.connector.execute_query_from_file(
                query_path,
                params={'run_id': self._get_run_id()}
            )
            df = self._normalize_columns(df)
            self._data_cache[query_name] = df
            print(f"    Loaded {len(df)} rows")
            print(f"    Columns: {df.columns.tolist()}")
            return df
        except Exception as e:
            logger.error(f"Error loading {query_name}: {e}")
            return pd.DataFrame()

    # -------------------------------------------------------------------------
    # Excel COM helpers
    # -------------------------------------------------------------------------
    def _excel_app(self):
        """Create an Excel COM Application with proper COM initialization."""
        import pythoncom
        import win32com.client as win32
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        return excel

    def _excel_refresh_all(self, file_path: Path) -> None:
        """Basic RefreshAll + async wait."""
        import pythoncom
        print(f"  Refreshing (Excel COM): {file_path.name}")
        self._make_writable(file_path)
        excel = self._excel_app()
        wb = None
        try:
            wb = excel.Workbooks.Open(str(file_path.resolve()), UpdateLinks=1, ReadOnly=False)
            wb.RefreshAll()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            wb.Save()
            wb.Close(SaveChanges=False)
            print("    Refresh complete")
        finally:
            wb = None
            excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    def _excel_calculate_and_refresh_all(self, file_path: Path) -> None:
        """Refresh all queries/links and calculate all formulas."""
        import pythoncom
        import win32com.client as win32
        print(f"  Refreshing and calculating (Excel COM): {file_path.name}")
        self._make_writable(file_path)
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.EnableEvents = False
            wb = excel.Workbooks.Open(str(file_path.resolve()), UpdateLinks=1, ReadOnly=False)
            wb.RefreshAll()
            wb.Application.Calculate()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            wb.Save()
            wb.Close(SaveChanges=False)
            print("    Refresh and calculation complete")
        finally:
            wb = None
            if excel:
                excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    def _excel_refresh_links_and_all(self, file_path: Path, dependency_path: Optional[Path] = None) -> None:
        """Refresh all workbook links, queries, and formulas in the workbook."""
        import pythoncom
        import win32com.client as win32
        print(f"  Refreshing all workbook links and calculating (Excel COM): {file_path.name}")
        self._make_writable(file_path)
        pythoncom.CoInitialize()
        excel = None
        wb = None
        dep_wb = None
        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            excel.EnableEvents = False

            if dependency_path and dependency_path.exists():
                print(f"    Opening dependency workbook in background: {dependency_path.name}")
                dep_wb = excel.Workbooks.Open(str(dependency_path.resolve()), UpdateLinks=0, ReadOnly=True)

            wb = excel.Workbooks.Open(str(file_path.resolve()), UpdateLinks=1, ReadOnly=False)
            xlExcelLinks = 1
            links = wb.LinkSources(xlExcelLinks)
            if links:
                wb.UpdateLinks = xlExcelLinks
            wb.RefreshAll()
            wb.Application.Calculate()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            wb.Save()
            wb.Close(SaveChanges=False)
            
            if dep_wb is not None:
                dep_wb.Close(SaveChanges=False)
            print("    Refresh, link update, and calculation complete")
        finally:
            dep_wb = None
            wb = None
            if excel:
                excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    # -------------------------------------------------------------------------
    # METRICS: writing data to sheets, link update, break links
    # -------------------------------------------------------------------------
    def _write_df_to_sheet_a_f(self, wb, sheet_name: str, df: pd.DataFrame):
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        max_row = ws.max_row
        if max_row >= 2:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.value = None

        if df.empty:
            print(f"    No data for {sheet_name}, leaving sheet empty.")
            return

        df_to_write = df.iloc[:, :6]
        for r_idx, row in enumerate(df_to_write.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        print(f"    Wrote {len(df_to_write)} rows into '{sheet_name}'")

    def _update_all_links_in_workbook(
        self,
        workbook_path: Path,
        vintage_gas_path: Optional[Path] = None,
        vintage_eua_path: Optional[Path] = None,
    ) -> bool:
        import pythoncom
        print(f"\nUpdating external links in: {workbook_path.name}")
        self._make_writable(workbook_path)
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            import win32com.client as win32
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            xlExcelLinks = 1
            wb = excel.Workbooks.Open(str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=False)
            links = wb.LinkSources(xlExcelLinks)
            if not links:
                print("  No external workbook links found.")
                wb.Close(SaveChanges=False)
                return True

            updated_count = 0
            for old_link in links:
                old_link_str = str(old_link)
                old_link_lower = old_link_str.lower()
                new_target = None
                label = None

                if vintage_gas_path and any(tok in old_link_lower for tok in ("gas", "henry")):
                    new_target = str(vintage_gas_path.resolve())
                    label = "Gas"
                elif vintage_eua_path and any(tok in old_link_lower for tok in ("eua", "carbon", "emission")):
                    new_target = str(vintage_eua_path.resolve())
                    label = "EUA"

                if not new_target:
                    continue

                try:
                    print(f"  Updating {label} link to: {Path(new_target).name}")
                    wb.ChangeLink(Name=old_link, NewName=new_target, Type=xlExcelLinks)
                    updated_count += 1
                except Exception as e:
                    print(f"  [WARNING] Failed to update link '{old_link_str}': {e}")

            if updated_count > 0:
                wb.Save()
                print(f"[OK] Updated {updated_count} link(s).")
            else:
                print("[INFO] No matching links changed.")
            wb.Close(SaveChanges=False)
            return True
        except Exception as e:
            print(f"[ERROR] Failed to update external links in {workbook_path.name}: {e}")
            return False
        finally:
            wb = None
            if excel:
                excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    def _break_all_links_in_workbook(self, workbook_path: Path) -> None:
        import pythoncom
        print(f"  Breaking external links in: {workbook_path.name}")
        self._make_writable(workbook_path)
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            import win32com.client as win32
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            xlExcelLinks = 1
            wb = excel.Workbooks.Open(str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=False)
            links = wb.LinkSources(xlExcelLinks)
            if links:
                broken_count = 0
                for link in links:
                    try:
                        wb.BreakLink(Name=link, Type=xlExcelLinks)
                        broken_count += 1
                    except Exception as e:
                        print(f"    [WARNING] Could not break link '{link}': {e}")
                print(f"    Successfully broke {broken_count} out of {len(links)} link(s).")
            else:
                print("    No external links to break.")
            wb.Save()
            wb.Close(SaveChanges=False)
        finally:
            wb = None
            if excel:
                excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    # -------------------------------------------------------------------------
    # METRICS: main generation workflow
    # -------------------------------------------------------------------------
    def generate_metrics(self) -> Dict[str, str]:
        from openpyxl import load_workbook
        print("\n" + "=" * 60)
        print("STEP 1: GENERATING EPA METRICS DELIVERABLE (UPDATED)")
        print("=" * 60)

        self._make_writable(self.metrics_static_file)
        self._make_writable(self.capacity_prices_file)

        if not self.metrics_static_file.exists():
            raise FileNotFoundError(f"Metrics data file not found: {self.metrics_static_file}")
        if not self.capacity_prices_file.exists():
            raise FileNotFoundError(f"Capacity prices file not found: {self.capacity_prices_file}")
        if not self.metrics_template_path.exists():
            raise FileNotFoundError(f"Metrics template not found: {self.metrics_template_path}")

        print("\nConnecting to Snowflake...")
        if not self.connector.connect():
            raise ConnectionError("Failed to connect to Snowflake")

        try:
            print(f"\n[1/5] Updating data in {self.metrics_static_file.name}")
            wb = load_workbook(self.metrics_static_file)
            
            df_price = self._load_query_data("epa_metrics_price")
            self._write_df_to_sheet_a_f(wb, "Prices", df_price)
            
            df_thermal = self._load_query_data("epa_metrics_thermal")
            self._write_df_to_sheet_a_f(wb, "Thermal_DB", df_thermal)
            
            df_wind_solar = self._load_query_data("epa_metrics_wind_solar")
            self._write_df_to_sheet_a_f(wb, "W&S_DB", df_wind_solar)
            
            df_nuclear = self._load_query_data("epa_nuclear_flex")
            self._write_df_to_sheet_a_f(wb, "Nuclear_DB", df_nuclear)
            
            wb.save(self.metrics_static_file)
            print(f"  Saved updated metrics data file: {self.metrics_static_file}")
        finally:
            try:
                self.connector.disconnect()
            finally:
                print("Disconnected from Snowflake")

        # Refresh database metrics
        print(f"\n[2/5] Refreshing {self.metrics_static_file.name}...")
        self._excel_calculate_and_refresh_all(self.metrics_static_file)

        # Refresh dependent capacity prices
        print(f"\n[3/5] Refreshing {self.capacity_prices_file.name}...")
        self._excel_refresh_links_and_all(
            self.capacity_prices_file, 
            dependency_path=self.metrics_static_file
        )

        # Template handling
        print("\n[4/5] Creating with-links Metrics template file...")
        timestamp = datetime.now().strftime("%Y%m")
        metrics_with_links = (
            self.output_dir_generated / f"EPA Metrics - Planning case {timestamp} (LINKS).xlsx"
        )
        shutil.copy2(self.metrics_template_path, metrics_with_links)
        self._make_writable(metrics_with_links)

        print("\nUpdating external links in with-links file...")
        _ = self._update_all_links_in_workbook(
            metrics_with_links,
            vintage_gas_path=self.vintage_gas_path,
            vintage_eua_path=self.vintage_eua_path,
        )

        print("\nRefreshing with-links Metrics file...")
        self._excel_calculate_and_refresh_all(metrics_with_links)

        # Create flattened flat output
        print("\n[5/5] Creating without-links Metrics file...")
        metrics_without_links = (
            self.output_dir_generated / f"EPA Metrics - Planning case {timestamp} (NO LINKS).xlsx"
        )
        shutil.copy2(metrics_with_links, metrics_without_links)
        self._make_writable(metrics_without_links)
        self._break_all_links_in_workbook(metrics_without_links)

        return {
            "metrics_data_file": str(self.metrics_static_file),
            "capacity_prices_file": str(self.capacity_prices_file),
            "metrics_with_links": str(metrics_with_links),
            "metrics_without_links": str(metrics_without_links),
        }

    # -------------------------------------------------------------------------
    # WHOLESALE: data extraction & transformations
    # -------------------------------------------------------------------------
    def _extract_hourly_prices(self) -> pd.DataFrame:
        print("  Extracting hourly prices from Snowflake...")
        query_path = self.query_manager.get_query_path("epa_hourly_prices")
        if not query_path.exists():
            raise FileNotFoundError(f"Query not found: {query_path}")
        df = self.connector.execute_query_from_file(
            query_path,
            params={"run_id": self._get_run_id()},
        )
        print(f"    Extracted {len(df)} rows, {len(df.columns)} columns")
        return df

    @staticmethod
    def _normalize_interval_headers(cols: List[str]) -> List[str]:
        return [c.replace("_", " ") if isinstance(c, str) else c for c in cols]

    def _excel_safe_scalar(self, x: Any) -> Any:
        if x is None or x is pd.NaT:
            return None
        try:
            if pd.isna(x):
                return None
        except Exception:
            pass
        if isinstance(x, pd.Timestamp):
            if x.hour == 0 and x.minute == 0 and x.second == 0 and x.microsecond == 0:
                s = x.strftime("%Y-%m-%d")
            else:
                s = x.strftime("%Y-%m-%d %H:%M:%S")
            return self._currency_pattern.sub("", s)
        if isinstance(x, (np.integer, np.floating, np.bool_)):
            return x.item()
        if isinstance(x, str):
            return self._currency_pattern.sub("", x)
        return x

    def _build_interval_matrix(self, df: pd.DataFrame, include_header: bool) -> list[list[Any]]:
        df2 = df.copy().astype(object)
        df2.columns = self._normalize_interval_headers([str(c) for c in df2.columns.tolist()])
        matrix: list[list[Any]] = []
        if include_header:
            matrix.append([self._excel_safe_scalar(c) for c in df2.columns.tolist()])
        for row in df2.values.tolist():
            matrix.append([self._excel_safe_scalar(v) for v in row])
        return matrix

    def _write_interval_raw_data_a4(self, df: pd.DataFrame, include_header: bool = True) -> None:
        import pythoncom
        print("  Writing data into interval workbook (raw data!A4) via Range.Value...")
        self._make_writable(self.interval_file_path)
        if not self.interval_file_path.exists():
            raise FileNotFoundError(f"Interval file not found: {self.interval_file_path}")

        matrix = self._build_interval_matrix(df, include_header=include_header)
        nrows = len(matrix)
        ncols = len(matrix[0]) if nrows else 0
        if nrows == 0 or ncols == 0:
            raise ValueError("No data to write to interval file.")

        excel = self._excel_app()
        wb = None
        try:
            wb = excel.Workbooks.Open(str(self.interval_file_path.resolve()), UpdateLinks=0, ReadOnly=False)
            ws = wb.Worksheets("raw data")
            start = ws.Range("A4")
            end = ws.Cells(start.Row + nrows - 1, start.Column + ncols - 1)
            target = ws.Range(start, end)
            target.ClearContents()
            target.Value = tuple(tuple(r) for r in matrix)
            wb.Save()
            wb.Close(SaveChanges=False)
            print(f"    Wrote {nrows - 1 if include_header else nrows} rows × {ncols} cols")
        finally:
            ws = None
            target = None
            wb = None
            excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    # -------------------------------------------------------------------------
    # WHOLESALE: vintage link updates
    # -------------------------------------------------------------------------
    @staticmethod
    def _iter_vintage_rules(
        gas: Optional[Path],
        eua: Optional[Path],
        coal: Optional[Path],
    ) -> Iterable[Tuple[str, Optional[Path], Tuple[str, ...]]]:
        return (
            ("Gas", gas, ("gas", "henry")),
            ("EUA", eua, ("eua", "carbon", "emission")),
            ("Coal", coal, ("coal", "ara", "co2")),
        )

    def _update_wholesale_vintage_links(self, workbook_path: Path) -> bool:
        import pythoncom
        if not self.vintage_gas_path and not self.vintage_eua_path and not self.vintage_coal_path:
            print("  No vintage paths specified - keeping template links")
            return True

        print("\n" + "=" * 60)
        print("Vintage Link Update (Wholesale Prices workbook)")
        print("=" * 60)
        self._make_writable(workbook_path)

        for label, p, _tokens in self._iter_vintage_rules(
            self.vintage_gas_path, self.vintage_eua_path, self.vintage_coal_path
        ):
            if p:
                if not p.exists():
                    print(f"[ERROR] {label} vintage file not found: {p}")
                    return False
                print(f"  {label} vintage: {p.name}")

        xlExcelLinks = 1
        excel = self._excel_app()
        wb = None
        try:
            wb = excel.Workbooks.Open(str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=False)
            links = wb.LinkSources(xlExcelLinks)
            if not links:
                print("  No external workbook links found (nothing to update)")
                wb.Close(SaveChanges=False)
                return True

            updated_count = 0
            for old_link in links:
                old_link_str = str(old_link)
                old_link_lower = old_link_str.lower()
                new_target: Optional[str] = None
                chosen_label: Optional[str] = None

                for label, vintage_path, tokens in self._iter_vintage_rules(
                    self.vintage_gas_path, self.vintage_eua_path, self.vintage_coal_path
                ):
                    if not vintage_path:
                        continue
                    if any(tok in old_link_lower for tok in tokens):
                        new_target = str(vintage_path.resolve())
                        chosen_label = label
                        break

                if not new_target:
                    continue

                try:
                    print(f"  Updating {chosen_label} link to: {Path(new_target).name}")
                    wb.ChangeLink(Name=old_link, NewName=new_target, Type=xlExcelLinks)
                    updated_count += 1
                except Exception as e:
                    print(f"  [WARNING] Failed to update link '{old_link_str}': {e}")

            if updated_count > 0:
                wb.Save()
                print(f"[OK] Updated {updated_count} link(s).")
            else:
                print("[INFO] No matching gas/eua/coal links found.")
            wb.Close(SaveChanges=False)
            return True
        except Exception as e:
            print(f"[ERROR] Failed to update vintage links: {e}")
            return False
        finally:
            wb = None
            excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    # -------------------------------------------------------------------------
    # WHOLESALE: without-links copy
    # -------------------------------------------------------------------------
    def _create_without_links_xlsb_break_links_only(self, source_xlsb: Path) -> Path:
        import pythoncom
        output_xlsb = source_xlsb.parent / f"{source_xlsb.stem}_without_links{source_xlsb.suffix}"
        print(f"  Creating without-links XLSB (break links only): {output_xlsb.name}")
        shutil.copy2(source_xlsb, output_xlsb)
        self._make_writable(output_xlsb)

        xlExcelLinks = 1
        excel = self._excel_app()
        wb = None
        try:
            wb = excel.Workbooks.Open(str(output_xlsb.resolve()), UpdateLinks=0, ReadOnly=False)
            links = wb.LinkSources(xlExcelLinks)
            if links:
                broken_count = 0
                for link in links:
                    try:
                        wb.BreakLink(Name=link, Type=xlExcelLinks)
                        broken_count += 1
                    except Exception as e:
                        print(f"    [WARNING] Could not break link '{link}': {e}")
                print(f"    Successfully broke {broken_count} out of {len(links)} link(s).")
            else:
                print("    No external links found to break.")
            
            # Save safely and exit
            wb.Save()
            wb.Close(SaveChanges=False)
            print("    Links broken process completed.")
            return output_xlsb
        except Exception as e:
            print(f"[ERROR] Exception occurred in saving/closing the stripped XLSB file: {e}")
            raise e
        finally:
            wb = None
            excel.Quit()
            excel = None
            pythoncom.CoUninitialize()

    # -------------------------------------------------------------------------
    # Public generation methods
    # -------------------------------------------------------------------------
    def generate_wholesale(self) -> Dict[str, str]:
        print("\n" + "=" * 60)
        print("STEP 2: GENERATING EPA WHOLESALE PRICES + INTERVAL DATA (XLSB)")
        print("=" * 60)

        self._make_writable(self.wholesale_template_path)
        self._make_writable(self.interval_file_path)

        if not self.wholesale_template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.wholesale_template_path}")
        if not self.interval_file_path.exists():
            raise FileNotFoundError(f"Interval file not found: {self.interval_file_path}")

        timestamp = datetime.now().strftime("%Y%m")
        with_links_xlsb = self.output_dir_generated / f"Wholesale prices - Planning Case {timestamp} (LINKS).xlsb"
        print(f"\n[1/8] Copying template to: {with_links_xlsb.name}")
        shutil.copy2(self.wholesale_template_path, with_links_xlsb)
        self._make_writable(with_links_xlsb)

        print("\n[2/8] Connecting to Snowflake...")
        if not self.connector.connect():
            raise ConnectionError("Failed to connect to Snowflake")
        try:
            print("\n[3/8] Extracting hourly prices...")
            df = self._extract_hourly_prices()
            print("\n[4/8] Populating interval file...")
            self._write_interval_raw_data_a4(df, include_header=True)
            print("\n[5/8] Refreshing interval file...")
            self._excel_refresh_all(self.interval_file_path)
            print("\n[6/8] Refreshing Wholesale Prices (xlsb) file...")
            self._excel_refresh_all(with_links_xlsb)
        finally:
            try:
                self.connector.disconnect()
            finally:
                print("Disconnected from Snowflake")

        print("\n[7/8] Updating vintage links (only gas/eua/coal if provided)...")
        ok = self._update_wholesale_vintage_links(with_links_xlsb)
        if not ok:
            raise RuntimeError("Vintage link update failed; aborting generation.")

        if self.vintage_gas_path or self.vintage_eua_path or self.vintage_coal_path:
            print("\n[8/8] Refreshing Wholesale Prices again...")
            self._excel_refresh_all(with_links_xlsb)
        else:
            print("\n[8/8] Skipping post-vintage refresh (no vintage paths specified).")

        print("\n[FINAL] Creating WITHOUT LINKS XLSB (break links only)...")
        without_links_xlsb = self._create_without_links_xlsb_break_links_only(with_links_xlsb)
        return {
            "Wholesale Prices (with links, xlsb)": str(with_links_xlsb),
            "Wholesale Prices (without links, xlsb)": str(without_links_xlsb),
        }

    def generate_all(self) -> Dict[str, str]:
        metrics_results = self.generate_metrics()
        wholesale_results = self.generate_wholesale()

        print("\nGenerated files:")
        for k, v in metrics_results.items():
            print(f"  {k}: {v}")
        for k, v in wholesale_results.items():
            print(f"  {k}: {v}")
            
        all_results = {}
        all_results.update(metrics_results)
        all_results.update(wholesale_results)
        return all_results


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate EPA Metrics and Wholesale Prices deliverables"
    )
    parser.add_argument("--run-id", help="Snowflake RUN_ID to use")
    parser.add_argument("--output-dir", help="Output directory")
    parser.add_argument("--template-dir", help="Template directory")
    parser.add_argument("--gas-vintage", help="Path to gas vintage workbook")
    parser.add_argument("--eua-vintage", help="Path to EUA vintage workbook")
    parser.add_argument("--coal-vintage", help="Path to coal vintage workbook")
    parser.add_argument("--dry-run", action="store_true", help="Show actions without running")
    args = parser.parse_args()

    if args.dry_run:
        print("DRY RUN - Would execute full generation pipeline.")
        return

    connector = SnowflakeConnector()
    vintage_gas_path = Path(args.gas_vintage) if args.gas_vintage else None
    vintage_eua_path = Path(args.eua_vintage) if args.eua_vintage else None
    vintage_coal_path = Path(args.coal_vintage) if args.coal_vintage else None

    generator = EPAGenerator(
        output_dir=args.output_dir,
        template_dir=args.template_dir,
        run_id=args.run_id,
        vintage_gas_path=str(vintage_gas_path) if vintage_gas_path else None,
        vintage_eua_path=str(vintage_eua_path) if vintage_eua_path else None,
        vintage_coal_path=str(vintage_coal_path) if vintage_coal_path else None,
        connector=connector,
    )
    generator.generate_all()


if __name__ == "__main__":
    main()
